import openpyxl, os

folder = r'c:\Users\Ny\Downloads\answersinSpecialEPSTOPIK'
files = os.listdir(folder)
for f in files:
    path = os.path.join(folder, f)
    try:
        wb = openpyxl.load_workbook(path)
        print(f'File: {repr(f)}')
        print(f'  Sheets: {wb.sheetnames}')
        for sh in wb.sheetnames:
            ws = wb[sh]
            print(f'  Sheet [{sh}]: {ws.max_row} rows x {ws.max_column} cols')
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                print(f'    {row}')
        print()
    except Exception as e:
        print(f'Error: {e}')
